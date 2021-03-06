# from openpyxl import Workbook
import openpyxl
import os.path

from openpyxl.workbook.workbook import Workbook

xlsx_file = './Coin_Trading_Bot.xlsx'
sheet = 'trading_bot_revenue'
# workbook 생성 (덮어쓰기)
# write_wb = Workbook()
if os.path.isfile(xlsx_file):
    write_wb = openpyxl.load_workbook(xlsx_file)
else:
    write_wb = Workbook()
    write_ws = write_wb.create_sheet(sheet)

# Set write worksheet
sheet_list = write_wb.sheetnames
flag = 0
for sheet_name in sheet_list :
    if sheet_name == 'Sheet':
        # 기본 생성 시트 제거
        write_wb.remove(write_wb[sheet_name])
    if sheet_name == sheet :
        flag = 1
        break

if flag == 0 :
    write_ws = write_wb.create_sheet('trading_bot_revenue')
else :
    write_ws = write_wb['trading_bot_revenue']

# 작업할 workbook 내 sheet 활성화
#write_ws = write_ws.active
write_ws['A1'] = '날짜'
write_ws['B1'] = '티커'     # 암호화폐 구분
write_ws['C1'] = '구분'     # 매도, 매수 구분
write_ws['D1'] = '매수금액'  # 매수 금액
write_ws['E1'] = 'MA15'    # 15 이동평균선
write_ws['F1'] = 'MA50'    # 50 이동평균선
write_ws['G1'] = '총액'     # 총액

# 행 단위 추가
# write_ws.append( [4, 5, 6] )


# 셀 단위 추가
# write_ws.cell(6, 6, '6행 6열')

# 저장
# write_wb.save('./Auto_Trading_Bot.xlsx')